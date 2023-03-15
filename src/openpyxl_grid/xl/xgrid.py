"""
MIT License

Copyright (c) 2023 Jianai Wang

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""
import os
import null as null
from datetime import *
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

from openpyxl_grid.xl.xoper import VType
from openpyxl_grid.xl.xlocation import XLocation
from openpyxl_grid.xl.xoper import OType
from openpyxl_grid.xl.xoper import XOperation

from openpyxl_grid.xl.xstyle import DefaultHeadFont
from openpyxl_grid.xl.xstyle import DefaultHeadFill
from openpyxl_grid.xl.xstyle import DefaultCellFont
from openpyxl_grid.xl.xstyle import DefaultBorder


class XParam(XLocation):
    name = null
    vType = VType.none
    oType = OType.none
    verbs = []
    vals = []
    customOper = null
    incrementIndex = 0
    isArrayFormula = False

    def __init__(self, name, v_type=VType.none, o_type=OType.none, row=0, col=0, vals=[], verbs=[], custom_oper=null):
        try:
            if (name == null):
                raise RuntimeError('xparam name can not be null')
        except RuntimeError as e:
            raise e
        super().__init__(row=row, col=col)
        self.name = name
        self.vType = v_type
        self.oType = o_type
        self.verbs = verbs
        self.vals = vals
        self.customOper = custom_oper
        self.isArrayFormula = OType.is_array_oper(o_type)

    def set_l_row_increment(self, i, is_inner_sheet, col_shift=0):
        for v in self.verbs:
            if (type(v) == XLocation):
                v.is_inner_sheet(is_inner_sheet)
                v.colShift = col_shift
                v.set_row_increment(i)
            elif (type(v) == XOperation):
                v.set_o_row_increment(i, is_inner_sheet, col_shift)
        self.set_row_increment(i)
        self.incrementIndex += i 

    def get_loc(self):
        return XLocation(sheet_name=self.sheetName, row=self.sRow, col=self.sCol, e_row=self.eRow, e_col=self.eCol)
    
    def get_val(self, increment_index=0, is_inner_sheet=False, col_shift=0):
        if (increment_index > 0):
            self.set_l_row_increment(i=increment_index, is_inner_sheet=is_inner_sheet, col_shift=col_shift)
        if (self.oType != OType.none or self.customOper != null):
            o = XOperation(v_type=self.vType, o_type=self.oType, verbs=self.verbs, custom_oper=self.customOper)
            return '={0}'.format(o.to_str())
        else:
            if (self.incrementIndex <= len(self.vals)):
                return self.vals[self.incrementIndex - 1] 
            else:
                return 0 if self.vType == VType.num or self.vType == VType.dateTime else ''


class XGrid:
    name = null
    params = []
    rows = 1

    def __init__(self, name, rows=1, params=[]):
        try:
            if (null == name):
                raise RuntimeError("xobj name can not be null")
        except RuntimeError as e:
            raise e
        try:
            if (rows < 1):
                raise RuntimeError('rows must be larger than 1')
        except RuntimeError as e:
            raise e
        if (params != null):
            self.params = params
        self.name = name
        self.rows = rows

    def __get_workbook(self, file_name):
        if (os.path.exists(file_name)):
            return load_workbook(file_name)
        else:
            wb = Workbook()
            wb.remove(wb['Sheet'])
            return wb
    
    def __get_sheet(self, wb, sheet_name):
        if sheet_name not in wb.sheetnames:
            return wb.create_sheet(sheet_name)
        else:
            return wb[sheet_name]

    def __get_write_sheet(self, wb, sheet_name, is_grid):
        if (sheet_name == null):
            return self.__get_sheet(wb, '{0}_{1}'.format('G' if is_grid else 'L', self.name))
        else:
            return self.__get_sheet(wb, sheet_name)


    def __get_label(self, row, col):
        return '{0}{1}'.format(get_column_letter(col), row)

    def __fit_col(self, ws, col, length):
        w = ws.column_dimensions[get_column_letter(col)].width
        if (w < length):
            ws.column_dimensions[get_column_letter(col)].width = length


    def clear_params(self):
        self.params = []

    def append_param(self, param):
        self.params.append(param)

    def write_xl(self, file_name, sheet_name=null, is_inner_sheet=False, col_shift=0, head_font=DefaultHeadFont
                 , head_fill=DefaultHeadFill, cell_font=DefaultCellFont, border=DefaultBorder):
        wb = self.__get_workbook(file_name)
        ws = self.__get_write_sheet(wb, sheet_name, True)
        c = 1
        for p in self.params:
            wc = ws.cell(row=1, column=c)
            wc.value = p.name
            wc.font = head_font
            wc.fill = head_fill
            wc.border = border
            self.__fit_col(ws, c, len(p.name) + 2)
            for r in range(self.rows):
                label = self.__get_label(row=r + 2, col=c)
                cv = p.get_val(1, is_inner_sheet, col_shift)
                if (cv != null):
                    if (p.isArrayFormula == False):
                        ws[label] = cv
                    else:
                        ws[label] = ArrayFormula(label, cv)
                if (p.vType == VType.dateTime and p.oType in (OType.addDays, OType.equal, OType.max, OType.replace
                                                              , OType.min, OType.maxIf, OType.minIf, OType.none)):
                    ws[label].number_format = 'yyyy-mm-dd'
                ws[label].font = cell_font
                ws[label].border = border
            c += 1
            p.set_l_row_increment(-1 - r, is_inner_sheet)
        wb.save(file_name)

